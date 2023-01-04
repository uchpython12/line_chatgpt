import requests
from pyngrok import ngrok     # pip install pyngrok
import openpyxl

def Line_設定Webhook(ngrokHTTP,auth_token):
    # 資料回傳 到 Line 的 https 伺服器
    # ngrokHTTP = ""
    message = {"endpoint": ngrokHTTP}
    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/channel/webhook/endpoint'
    t1 = requests.put(url, json=message, headers=hed)  # 把資料HTTP POST送出去
    print(t1)
    return t1

def Line_取得用戶的資訊(userId,auth_token):
    # 資料回傳 到 Line 的 https 伺服器
    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/profile/'+userId
    t1 = requests.get(url, headers=hed)  # 把資料HTTP POST送出去
    print(t1)
    if t1.status_code == 200:
        return  t1.text
    return t1

def ngrok_啟動(port=8888,protocol="http"):

    ngrok_關閉()
    ssh_tunnel = ngrok.connect(port,protocol)  # 8888, "http")        # 開啟 localhost:8888 的 HTTP 連線
    publicIP=""
    tunnels = ngrok.get_tunnels()                   # 取得所有的tunnel
    for x in tunnels:
        if x.config["addr"].lower()=="http://localhost:"+str(port):
            if x.proto.lower() == "https":
                publicIP=x.data["public_url"]
                print(publicIP)         # 取得第二個tunnel的public_url
                break
    return publicIP

def ngrok_持續執行():
    ngrok_process = ngrok.get_ngrok_process()
    try:
        # Block until CTRL-C or some other terminating event
        print(" CTRL-C 關閉程式")
        ngrok_process.proc.wait()
    except KeyboardInterrupt:
        ngrok_關閉()


def ngrok_關閉():
    print(" Shutting down server.")
    tunnels = ngrok.get_tunnels()                   # 取得所有的tunnel
    for x in tunnels:
        ngrok.disconnect(x.public_url)  # 取消連線
    ngrok.kill()

import uuid
def UUID_產生器():
    return  str(uuid.uuid4())


def Line_回送文字(replyToken="",text="",userId=""):
    if text=="":
        str1="你的 User Id: " + userId + "\n 傳過來的文字 Text:" + text
    else:
        str1=text

    message = {
        "replyToken": replyToken,
        "messages": [
            {
                "type": "text",
                "text": str1
            }
        ]
    }
    return Line_回送(message)

def Line_回送(auth_token,message=""):

    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/message/reply'
    response = requests.post(url, json=message, headers=hed)
    return response

def Line_廣播推送(auth_token,toID="",str1="hello"):
    message={
            "to": toID,
            "messages":[
                {
                    "type":"text",
                    "text":str1
                }
            ]
        }

    UUID=UUID_產生器()
    hed = {'Content-Type': 'application/json',
           'Authorization': 'Bearer ' + auth_token,
           'X-Line-Retry-Key':UUID}
    url = 'https://api.line.me/v2/bot/message/push'
    response = requests.post(url, json=message, headers=hed)
    print(response)

def Line_廣播到所有用戶(auth_token,str1="hello"):
    message={
            "messages":[
                {
                    "type":"text",
                    "text":str1
                }
            ]
        }


    UUID=UUID_產生器()
    hed = {'Content-Type': 'application/json',
           'Authorization': 'Bearer ' + auth_token,
           'X-Line-Retry-Key':UUID}
    url = 'https://api.line.me/v2/bot/message/broadcast'
    response = requests.post(url, json=message, headers=hed)
    print(response)
def Line_讀取設定檔Excel(filename):
    global auth_token,YouruserID,userId,groupID,接收人的id
    wb = openpyxl.load_workbook(filename) # 'line.xlsx')  # 讀取檔案
    # 方法一打開第一個 工作表單
    sheetSetup = wb["setup"]         # 打開一個工作欄
    auth_token=sheetSetup.cell(row=2, column=1).value
    接收人的id=sheetSetup.cell(row=2, column=2).value
    openai_key=sheetSetup.cell(row=2, column=3).value
    return auth_token,接收人的id,openai_key

def ngrok自動啟動加自動上傳Webhook網址():
    ngrokHTTP=ngrok_啟動()
    # auth_token,接收人的id=auth_token,userId
    auth_token,接收人的id,openai_key=Line_讀取設定檔Excel('line.xlsx')
    t1=Line_設定Webhook(ngrokHTTP,auth_token)

    # print(ngrokHTTP)
    print("ngrokhttp 自動啟動 "+str(t1))

    #將ngrokHTTP 回傳到 line後台

    # 資料回傳 到 Line 的 https 伺服器

    message={"endpoint":ngrokHTTP}
    hed = {'Authorization': 'Bearer ' + auth_token}
    url = 'https://api.line.me/v2/bot/channel/webhook/endpoint'
    t1=requests.put(url, json=message, headers=hed)  # 把資料HTTP POST送出去

    print("Webhook網址 資料回傳 到 Line 的 https 伺服器"+ str(t1))
    ngrok_持續執行()